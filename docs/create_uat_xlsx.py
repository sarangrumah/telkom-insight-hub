"""
Generate UAT_SCENARIOS_SIGNED.xlsx
528 scenarios across 24 sheets, all with sign-off boxes.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
from openpyxl.utils import get_column_letter
import os

OUTPUT_DIR = os.path.dirname(__file__)

wb = Workbook()

# ── Styles ─────────────────────────────────────────────────────────────────
RED_FILL = PatternFill(start_color="E11A27", end_color="E11A27", fill_type="solid")
DARK_FILL = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
LIGHT_GRAY_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
GREEN_FILL = PatternFill(start_color="2ECC71", end_color="2ECC71", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
BLUE_FILL = PatternFill(start_color="0096D6", end_color="0096D6", fill_type="solid")
SIGNOFF_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

WHITE_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
RED_FONT = Font(name="Calibri", size=10, bold=True, color="E11A27")
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
NORMAL_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(name="Calibri", size=10, bold=True)
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="E11A27")
SIGNOFF_FONT = Font(name="Calibri", size=10, bold=True, color="333333")

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_WRAP = Alignment(horizontal='left', vertical='top', wrap_text=True)
CENTER_WRAP = Alignment(horizontal='center', vertical='center', wrap_text=True)

COLUMNS = ["No", "Scenario ID", "Test Scenario", "Pre-condition", "Test Steps", "Expected Result", "Priority", "Status", "Tester", "Date", "Notes"]
COL_WIDTHS = [5, 14, 35, 25, 40, 30, 10, 10, 12, 12, 20]

def create_sheet(ws, sheet_title, module_name, scenarios, start_id):
    """Create a UAT sheet with scenarios and sign-off box."""
    ws.title = sheet_title

    # Title row
    ws.merge_cells('A1:K1')
    cell = ws['A1']
    cell.value = f"UAT SCENARIOS — {module_name}"
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:K2')
    cell = ws['A2']
    cell.value = f"Telkom Insight Hub — Kementerian Komunikasi dan Digital | Module: {module_name}"
    cell.font = Font(name="Calibri", size=10, color="666666")
    cell.alignment = Alignment(horizontal='center')

    # Header row (row 4)
    header_row = 4
    for col_idx, (header, width) in enumerate(zip(COLUMNS, COL_WIDTHS), 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = RED_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows
    for i, scenario in enumerate(scenarios):
        row = header_row + 1 + i
        sc_id = f"UAT-{sheet_title}-{start_id + i:03d}"
        row_data = [
            i + 1,
            sc_id,
            scenario[0],  # Test Scenario
            scenario[1],  # Pre-condition
            scenario[2],  # Test Steps
            scenario[3],  # Expected Result
            scenario[4],  # Priority
            "",  # Status (to be filled)
            "",  # Tester
            "",  # Date
            "",  # Notes
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = NORMAL_FONT
            cell.alignment = LEFT_WRAP if col_idx in (3, 4, 5, 6) else CENTER
            cell.border = THIN_BORDER
            if i % 2 == 1:
                cell.fill = LIGHT_GRAY_FILL

    # Summary row
    summary_row = header_row + len(scenarios) + 2
    ws.merge_cells(f'A{summary_row}:B{summary_row}')
    cell = ws.cell(row=summary_row, column=1, value="Total Scenarios:")
    cell.font = BOLD_FONT
    cell = ws.cell(row=summary_row, column=3, value=len(scenarios))
    cell.font = BOLD_FONT

    # Status summary
    sr = summary_row + 1
    for label, fill in [("Pass", GREEN_FILL), ("Fail", RED_FILL), ("Blocked", YELLOW_FILL), ("Not Tested", BLUE_FILL)]:
        ws.cell(row=sr, column=1, value=label).font = BOLD_FONT
        ws.cell(row=sr, column=1).fill = fill
        ws.cell(row=sr, column=1).font = WHITE_FONT
        ws.cell(row=sr, column=2, value="=COUNTIF(H5:H{},\"{}\")".format(header_row + len(scenarios), label)).font = NORMAL_FONT
        sr += 1

    # Sign-off box
    signoff_start = sr + 2
    ws.merge_cells(f'A{signoff_start}:K{signoff_start}')
    cell = ws.cell(row=signoff_start, column=1, value="SIGN-OFF / PERSETUJUAN")
    cell.font = Font(name="Calibri", size=12, bold=True, color="E11A27")
    cell.alignment = Alignment(horizontal='center')

    signoff_headers = ["Peran", "Nama", "Jabatan", "Tanda Tangan", "Tanggal"]
    signoff_cols = [1, 3, 5, 7, 9]
    signoff_widths = [2, 2, 2, 2, 2]

    sr = signoff_start + 1
    for ci, (col, header) in enumerate(zip(signoff_cols, signoff_headers)):
        end_col = col + 1
        ws.merge_cells(start_row=sr, start_column=col, end_row=sr, end_column=end_col)
        cell = ws.cell(row=sr, column=col, value=header)
        cell.font = WHITE_FONT
        cell.fill = DARK_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        ws.cell(row=sr, column=end_col).border = THIN_BORDER

    signoff_roles = [
        "Disusun oleh (Prepared by)",
        "QA Lead (Reviewed by)",
        "Project Manager (Approved by)",
        "Stakeholder (Accepted by)",
    ]

    for ri, role in enumerate(signoff_roles):
        row = sr + 1 + ri
        for ci, col in enumerate(signoff_cols):
            end_col = col + 1
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=end_col)
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.alignment = CENTER
            cell.fill = SIGNOFF_FILL
            ws.cell(row=row, column=end_col).border = THIN_BORDER
            if ci == 0:
                cell.value = role
                cell.font = SIGNOFF_FONT

        ws.row_dimensions[row].height = 40  # Space for signature

    return len(scenarios)


# ═══════════════════════════════════════════════════════════════════════════
# Define all 24 modules with scenarios
# ═══════════════════════════════════════════════════════════════════════════

ALL_MODULES = {
    # ── Module 1: Authentication ──────────────────────────────────────────
    "AUTH": ("Authentication & Login", [
        ("Login with valid credentials", "User registered in system", "1. Navigate to login page\n2. Enter valid email\n3. Enter valid password\n4. Click Login", "User logged in, redirected to dashboard, JWT token received", "High"),
        ("Login with invalid password", "User registered in system", "1. Navigate to login page\n2. Enter valid email\n3. Enter wrong password\n4. Click Login", "Error message displayed, login rejected", "High"),
        ("Login with non-existent email", "No account with test email", "1. Navigate to login page\n2. Enter non-existent email\n3. Enter any password\n4. Click Login", "Error message: account not found", "High"),
        ("Login with empty fields", "Login page open", "1. Leave email empty\n2. Leave password empty\n3. Click Login", "Validation error shown for required fields", "High"),
        ("Login via e-Telekomunikasi", "User has e-Telekomunikasi account", "1. Click 'Login via e-Telekomunikasi'\n2. Enter e-Telekomunikasi credentials\n3. Authenticate", "User logged in via external auth, profile synced", "High"),
        ("JWT token refresh", "User logged in, token about to expire", "1. Wait for token expiration\n2. Make API request\n3. System auto-refreshes token", "New access token issued, session continues", "High"),
        ("Logout", "User logged in", "1. Click Logout button\n2. Confirm logout", "Session cleared, cookies removed, redirected to login", "High"),
        ("Rate limiting on login", "Login page open", "1. Attempt login 100+ times in 15 min\n2. Use wrong credentials each time", "Rate limit triggered, 429 response returned", "Medium"),
        ("Remember me / persistent session", "Login page open", "1. Login with valid credentials\n2. Close browser\n3. Reopen and navigate to dashboard", "Refresh token restores session within 30 days", "Medium"),
        ("Password validation on login", "Login page open", "1. Enter email with spaces\n2. Enter password < 8 chars", "Client-side validation errors shown", "Low"),
        ("Session expiry handling", "User logged in, session expired", "1. Wait for access token to expire\n2. Attempt navigation\n3. Refresh token also expired", "User redirected to login page with session expired message", "High"),
        ("Concurrent session handling", "User logged in on device A", "1. Login on device B with same credentials\n2. Verify both sessions active", "Both sessions remain active (multi-device support)", "Medium"),
        ("CSRF protection verification", "Login page open", "1. Attempt cross-site login request\n2. Verify CSRF token requirement", "Cross-site request blocked", "High"),
        ("Login audit trail", "Admin has access to audit logs", "1. Login as user\n2. Check audit_logs table\n3. Verify login event recorded", "Login event logged with timestamp, IP, user agent", "Medium"),
        ("Email check availability", "Registration page open", "1. Enter existing email\n2. Wait for validation", "Email marked as unavailable", "Medium"),
        ("Email check for new email", "Registration page open", "1. Enter new unique email\n2. Wait for validation", "Email marked as available", "Medium"),
        ("Login with SQL injection attempt", "Login page open", "1. Enter SQL injection in email field\n2. Click Login", "Input sanitized, login fails safely", "High"),
        ("Login with XSS in email field", "Login page open", "1. Enter <script> tag in email\n2. Click Login", "XSS attempt sanitized, no script execution", "High"),
        ("Token tampering detection", "User logged in", "1. Manually modify JWT payload\n2. Make API request", "Tampered token rejected, 401 returned", "High"),
        ("Refresh token reuse prevention", "Refresh token used", "1. Use refresh token\n2. Try to reuse same token", "Reused token rejected (rotation enforced)", "High"),
        ("Login page accessibility", "Login page open", "1. Navigate with keyboard only\n2. Use screen reader", "All form elements accessible via keyboard", "Low"),
        ("Password masking on input", "Login page open", "1. Type password\n2. Verify masked display", "Password characters masked", "Low"),
        ("Login redirect after session timeout", "Session expired", "1. Session times out\n2. User tries action\n3. Redirect to login\n4. Login\n5. Redirect back", "User returned to original page after login", "Medium"),
        ("Login form autofill", "Browser with saved credentials", "1. Open login page\n2. Verify browser autofill works", "Login credentials auto-filled from browser", "Low"),
        ("Login error message clarity", "Login failed", "1. Fail login\n2. Read error message", "Error message helpful without revealing user existence", "Medium"),
    ]),

    # ── Module 2: Registration ────────────────────────────────────────────
    "REG": ("Public Registration", [
        ("Register with complete data", "Registration page open", "1. Fill all required fields\n2. Upload all documents\n3. Fill PIC data\n4. Submit", "Registration successful, status: submitted", "High"),
        ("Register with missing required fields", "Registration page open", "1. Leave company name empty\n2. Submit", "Validation error for required fields", "High"),
        ("Upload NIB document", "Registration form open", "1. Click upload NIB\n2. Select PDF file\n3. Verify upload", "NIB document uploaded and previewed", "High"),
        ("Upload NPWP document", "Registration form open", "1. Click upload NPWP\n2. Select PDF file\n3. Verify upload", "NPWP document uploaded and previewed", "High"),
        ("Upload Akta Pendirian", "Registration form open", "1. Click upload Akta\n2. Select PDF file", "Akta document uploaded successfully", "High"),
        ("Upload KTP PIC", "Registration form open", "1. Click upload KTP\n2. Select image file", "KTP uploaded and previewed", "High"),
        ("Upload invalid file type", "Registration form open", "1. Try to upload .exe file", "File type rejected with error message", "High"),
        ("Upload oversized file (>10MB)", "Registration form open", "1. Try to upload file > 10MB", "File size limit error displayed", "Medium"),
        ("Fill PIC information", "Registration form open", "1. Enter PIC name\n2. Enter PIC phone\n3. Enter PIC email\n4. Enter PIC position", "PIC data saved successfully", "High"),
        ("Register with existing email", "Email already registered", "1. Enter existing email\n2. Submit", "Error: email already registered", "High"),
        ("Register with valid NIB format", "Registration form open", "1. Enter valid NIB number\n2. Verify format", "NIB accepted", "Medium"),
        ("Register with invalid NIB", "Registration form open", "1. Enter invalid NIB format\n2. Verify", "Validation error shown", "Medium"),
        ("View registration confirmation", "Registration submitted", "1. Complete registration\n2. View confirmation page", "Confirmation with status 'submitted' shown", "High"),
        ("Multi-step form navigation", "Registration form open", "1. Fill step 1\n2. Go to step 2\n3. Go back to step 1\n4. Verify data preserved", "Data persisted across form steps", "Medium"),
        ("Form auto-save", "Registration form partially filled", "1. Fill some fields\n2. Navigate away\n3. Return to form", "Previously entered data still present", "Low"),
        ("Location selector - Province", "Registration form open", "1. Open province dropdown\n2. Select 'DKI Jakarta'", "Province selected, kabupaten list updates", "High"),
        ("Location selector - Kabupaten", "Province selected", "1. Open kabupaten dropdown\n2. Select kabupaten", "Kabupaten selected successfully", "High"),
        ("Phone number validation", "Registration form open", "1. Enter invalid phone format\n2. Verify validation", "Phone number validation error shown", "Medium"),
        ("NPWP format validation", "Registration form open", "1. Enter invalid NPWP\n2. Verify validation", "NPWP validation error shown", "Medium"),
        ("Registration form responsive", "Mobile device/viewport", "1. Open registration on mobile\n2. Verify all fields accessible", "Form renders correctly on mobile", "Low"),
        ("Registration with special chars in name", "Registration form open", "1. Enter company name with special characters\n2. Submit", "Special characters handled correctly", "Medium"),
        ("Registration form step indicator", "Registration form open", "1. View step indicator\n2. Progress through steps", "Step indicator shows current progress", "Low"),
        ("Registration terms acceptance", "Registration form last step", "1. Try submit without accepting terms\n2. Accept terms\n3. Submit", "Terms must be accepted before submission", "Medium"),
        ("Registration success redirect", "Registration completed", "1. Complete registration\n2. Verify redirect", "Redirected to confirmation/login page", "Medium"),
        ("Duplicate company registration", "Company already registered", "1. Enter same company NIB\n2. Submit", "Duplicate company detection warning", "High"),
        ("Registration field length limits", "Registration form open", "1. Enter very long text in fields\n2. Verify limits", "Character limits enforced on all fields", "Medium"),
        ("Registration captcha", "Registration form", "1. Verify captcha/bot protection\n2. Complete captcha", "Bot protection active on registration", "Medium"),
        ("Registration loading state", "Form submitting", "1. Click submit\n2. Observe loading state", "Submit button disabled, loading indicator shown", "Low"),
    ]),

    # ── Module 3: Company Verification ────────────────────────────────────
    "COMP": ("Company Verification", [
        ("View pending companies", "Admin logged in", "1. Navigate to Company Management\n2. Filter by status 'submitted'", "List of pending companies displayed", "High"),
        ("Review company details", "Admin on Company Management", "1. Click on a company\n2. View all submitted data", "Company details and documents displayed", "High"),
        ("Approve company", "Company in 'in_review' status", "1. Review documents\n2. Click 'Verify'\n3. Confirm", "Status changed to 'verified', user notified", "High"),
        ("Reject company", "Company in 'in_review' status", "1. Review documents\n2. Click 'Reject'\n3. Enter reason\n4. Confirm", "Status changed to 'rejected' with reason", "High"),
        ("Request correction", "Company in 'in_review' status", "1. Review documents\n2. Click 'Needs Correction'\n3. Add correction notes", "Status changed to 'needs_correction', notes saved", "High"),
        ("View correction notes", "Pelaku usaha logged in", "1. Check company status\n2. View correction notes", "Correction notes from admin displayed", "High"),
        ("Re-submit after correction", "Company status 'needs_correction'", "1. Fix identified issues\n2. Re-upload documents\n3. Re-submit", "Status changed back to 'submitted'", "High"),
        ("Filter companies by status", "Admin on Company Management", "1. Select status filter\n2. Apply filter", "Companies filtered by selected status", "Medium"),
        ("Search company by name", "Admin on Company Management", "1. Enter company name in search\n2. Submit search", "Matching companies displayed", "Medium"),
        ("View uploaded documents", "Admin reviewing company", "1. Click on document link\n2. View/download document", "Document opens in viewer/downloads", "High"),
        ("Bulk verification", "Multiple companies pending", "1. Select multiple companies\n2. Click bulk verify", "All selected companies verified", "Medium"),
        ("Verification audit trail", "Company verified", "1. Check audit logs\n2. Filter by company", "Verification action logged with admin name and timestamp", "Medium"),
        ("Company profile completion check", "Pelaku usaha logged in", "1. Navigate to profile\n2. Check completion status", "Profile completion percentage shown", "Medium"),
        ("PIC document review", "Admin reviewing company", "1. Navigate to PIC section\n2. Review PIC documents", "PIC documents displayed for review", "High"),
        ("Change company status workflow", "Company in various statuses", "1. Verify status transitions\n2. submitted > in_review > verified\n3. submitted > in_review > rejected\n4. submitted > in_review > needs_correction > submitted", "All valid transitions work correctly", "High"),
        ("Prevent invalid status transition", "Company in verified status", "1. Try to move to 'submitted' status", "Invalid transition blocked", "Medium"),
        ("Company verification notification", "Company just verified", "1. Login as pelaku usaha\n2. Check notifications", "Verification status notification received", "Medium"),
        ("Verification dashboard metrics", "Admin on Company Management", "1. View dashboard metrics\n2. Check counts per status", "Accurate counts for each status displayed", "Low"),
        ("Export company list", "Admin on Company Management", "1. Click Export\n2. Download file", "Company list exported to Excel", "Low"),
        ("View verification history", "Company with multiple reviews", "1. Open company detail\n2. View history tab", "All review actions displayed chronologically", "Medium"),
        ("Concurrent verification conflict", "Two admins reviewing same company", "1. Admin A starts review\n2. Admin B starts review\n3. Both try to verify", "Conflict handled, only one verification succeeds", "Medium"),
        ("Company document expiry check", "Company documents uploaded", "1. Check document validity dates\n2. Flag expired documents", "Expired documents highlighted", "Low"),
        ("Company search by NIB", "Admin on Company Management", "1. Enter NIB in search\n2. Submit", "Company with matching NIB displayed", "Medium"),
        ("Company search by NPWP", "Admin on Company Management", "1. Enter NPWP in search\n2. Submit", "Company with matching NPWP displayed", "Medium"),
        ("Company detail print view", "Company detail open", "1. Click Print\n2. Verify layout", "Print-friendly view generated", "Low"),
        ("Company list export", "Company management page", "1. Click export\n2. Download", "Company list exported to Excel", "Low"),
        ("Verification deadline tracking", "Companies pending review", "1. View pending duration\n2. Check SLA", "Duration since submission displayed", "Medium"),
        ("Company document re-upload", "Needs correction status", "1. Re-upload corrected document\n2. Save", "New document replaces old one", "High"),
        ("Admin notes persistence", "Admin adds notes", "1. Add verification notes\n2. Save\n3. Reload page", "Notes persisted and visible", "Medium"),
        ("Company count by status dashboard", "Companies in various statuses", "1. View status counts\n2. Verify accuracy", "Correct count per status displayed", "Medium"),
        ("Company sorting", "Company list displayed", "1. Sort by name\n2. Sort by date\n3. Sort by status", "Company list sorts correctly by all columns", "Medium"),
        ("Company pagination", "Many companies registered", "1. Navigate pages\n2. Verify data consistency", "Company list paginates correctly", "Medium"),
    ]),

    # ── Module 4: Telekom Data CRUD ───────────────────────────────────────
    "TDATA": ("Telekom Data Management", [
        ("View telekom data list", "User logged in", "1. Navigate to Data Management\n2. View data table", "Paginated list of telekom data displayed", "High"),
        ("Create new telekom record", "User with create permission", "1. Click 'Add New'\n2. Fill all fields\n3. Save", "New record created successfully", "High"),
        ("Edit existing record", "Record exists in system", "1. Click Edit on a record\n2. Modify fields\n3. Save", "Record updated successfully", "High"),
        ("Delete record", "Record exists in system", "1. Click Delete on a record\n2. Confirm deletion", "Record deleted, removed from list", "High"),
        ("View record detail", "Record exists in system", "1. Click on record\n2. View detail page", "Full record details displayed", "High"),
        ("Search by company name", "Records exist", "1. Enter company name in search\n2. Submit", "Matching records displayed", "High"),
        ("Filter by service type", "Records exist", "1. Select service type filter\n2. Apply", "Records filtered by service type", "High"),
        ("Filter by status", "Records exist", "1. Select status filter\n2. Apply", "Records filtered by status", "Medium"),
        ("Filter by province", "Records exist", "1. Select province filter\n2. Apply", "Records filtered by province", "High"),
        ("Filter by kabupaten", "Province selected", "1. Select kabupaten filter\n2. Apply", "Records filtered by kabupaten", "Medium"),
        ("Filter by date range", "Records exist", "1. Set start date\n2. Set end date\n3. Apply", "Records within date range displayed", "Medium"),
        ("Pagination - next page", "More than 1 page of data", "1. Click next page\n2. Verify data loads", "Next page of data displayed", "High"),
        ("Pagination - page size change", "Multiple records", "1. Change page size to 50\n2. Verify", "50 records per page displayed", "Medium"),
        ("Sort by column", "Records displayed", "1. Click column header\n2. Verify sort order", "Data sorted by selected column", "Medium"),
        ("Sort ascending/descending toggle", "Sorted data", "1. Click same column header\n2. Verify sort reversal", "Sort order toggled", "Medium"),
        ("Excel import - valid file", "User with import permission", "1. Click Import\n2. Upload valid Excel\n3. Map columns\n4. Confirm", "Data imported successfully", "High"),
        ("Excel import - column mapping", "Import dialog open", "1. Upload Excel\n2. Map source to target columns\n3. Preview mapping", "Columns correctly mapped", "High"),
        ("Excel import - preview data", "Columns mapped", "1. Review data preview\n2. Verify data accuracy", "Preview shows correct data", "High"),
        ("Excel import - invalid data", "Import dialog open", "1. Upload Excel with invalid data\n2. Attempt import", "Validation errors shown, import blocked", "Medium"),
        ("Excel export", "Records displayed", "1. Click Export\n2. Download file", "Excel file downloaded with current data", "High"),
        ("Multi-filter combination", "Records exist", "1. Set province filter\n2. Set service type filter\n3. Set status filter\n4. Apply all", "Records matching all criteria displayed", "Medium"),
        ("Clear all filters", "Filters applied", "1. Click 'Clear Filters'\n2. Verify reset", "All filters cleared, full data displayed", "Medium"),
        ("Empty state - no results", "Filters applied with no matches", "1. Apply filter with no matching data", "Empty state message displayed", "Low"),
        ("Record validation - required fields", "Create/Edit form open", "1. Leave required fields empty\n2. Submit", "Validation errors for required fields", "High"),
        ("Location-based data view", "Records with location data", "1. Apply location filter\n2. View results", "Only records from selected location shown", "Medium"),
        ("Batch delete records", "Multiple records selected", "1. Select multiple records\n2. Click batch delete\n3. Confirm", "Selected records deleted", "Low"),
        ("Data table column resize", "Data table displayed", "1. Drag column border\n2. Resize column", "Column width adjustable", "Low"),
        ("Data table column visibility", "Data table displayed", "1. Toggle column visibility\n2. Verify hidden/shown", "Columns can be shown/hidden", "Low"),
        ("Data record history/audit", "Record modified", "1. View record\n2. Check modification history", "All changes logged with timestamp and user", "Medium"),
        ("Quick search from table", "Data table displayed", "1. Use quick search box\n2. Type keyword", "Table filters in real-time", "Medium"),
        ("Data validation - numeric fields", "Create form open", "1. Enter text in numeric field\n2. Verify validation", "Numeric validation enforced", "Medium"),
        ("Data validation - date fields", "Create form open", "1. Enter invalid date\n2. Verify validation", "Date validation enforced", "Medium"),
        ("Copy record data", "Record displayed", "1. Click Copy\n2. Paste in new form", "Record data copied for new entry", "Low"),
        ("Record count per service type", "Data management page", "1. View service type tabs\n2. Check counts", "Correct record count per service type", "Medium"),
        ("Data table print view", "Data table displayed", "1. Click Print\n2. Verify print layout", "Print-friendly table layout generated", "Low"),
        ("Data record bookmark", "Record detail view", "1. Bookmark record URL\n2. Reopen from bookmark", "Record detail loads from bookmark URL", "Low"),
    ]),

    # ── Module 5: Service Pages ───────────────────────────────────────────
    "SVC": ("Service Type Pages", [
        ("View Jasa page", "Public user", "1. Navigate to /services/jasa\n2. View data", "Jasa service data displayed correctly", "High"),
        ("View Jaringan page", "Public user", "1. Navigate to /services/jaringan\n2. View data", "Jaringan service data displayed", "High"),
        ("View Telsus page", "Public user", "1. Navigate to /services/telsus\n2. View data", "Telsus service data displayed", "High"),
        ("View ISR page", "Public user", "1. Navigate to /services/isr\n2. View data", "ISR service data displayed", "High"),
        ("View SKLO page", "Public user", "1. Navigate to /services/sklo\n2. View data", "SKLO service data displayed", "High"),
        ("View LKO page", "Public user", "1. Navigate to /services/lko\n2. View data", "LKO service data displayed", "High"),
        ("View Penomoran page", "Public user", "1. Navigate to /services/penomoran\n2. View data", "Penomoran service data displayed", "High"),
        ("View Tarif page", "Public user", "1. Navigate to /services/tarif\n2. View data", "Tarif service data displayed", "High"),
        ("Search within service page", "Service page open", "1. Use search functionality\n2. Enter keyword", "Relevant results displayed within service type", "Medium"),
        ("Service page responsive", "Mobile viewport", "1. Open service page on mobile\n2. Verify layout", "Page renders correctly on mobile", "Low"),
        ("Navigate between services", "Service page open", "1. Click navigation to different service\n2. Verify data updates", "Correct service data loaded", "Medium"),
        ("Service data pagination", "Service page with many records", "1. Navigate to next page\n2. Verify data loads", "Pagination works on service pages", "Medium"),
        ("Service detail link to full record", "Service page open", "1. Click on a record\n2. Verify navigation", "Navigates to full record detail page", "Medium"),
        ("Public data search from homepage", "Homepage open", "1. Use search from homepage\n2. Enter company or license number", "Search results displayed from public data", "High"),
        ("Public data view - pagination", "Public data view open", "1. Navigate pages\n2. Verify data consistency", "Pagination works on public data view", "Medium"),
        ("Public data view - filter", "Public data view open", "1. Apply filters\n2. Verify results", "Data filtered correctly in public view", "Medium"),
        ("Telekom data detail view", "Record exists", "1. Navigate to /telekom/:id\n2. View full details", "Complete record details with related data", "High"),
        ("Service type icon/badge display", "Service page open", "1. Verify service type badges\n2. Check color coding", "Correct badges and colors per service type", "Low"),
        ("Service page loading skeleton", "Service page loading", "1. Navigate to service page\n2. Observe loading state", "Skeleton/loading indicator shown", "Low"),
        ("Service data export per type", "Service page open", "1. Click export on service page\n2. Download", "Service-specific data exported", "Medium"),
        ("Cross-service search", "Search page open", "1. Search keyword across all services\n2. View results", "Results from multiple service types shown", "Medium"),
        ("Service statistics summary", "Service page open", "1. View summary statistics\n2. Verify totals", "Correct statistics for service type", "Medium"),
        ("Service page breadcrumb", "Service detail page", "1. View breadcrumb navigation\n2. Click parent link", "Breadcrumb shows: Home > Services > [Type]", "Low"),
        ("Service data card view toggle", "Service page open", "1. Toggle between table and card view\n2. Verify", "Both view modes work correctly", "Low"),
        ("Service page URL routing", "Direct URL access", "1. Navigate to /services/jasa directly\n2. Verify page loads", "Service page loads from direct URL", "Medium"),
        ("Service type metadata display", "Service page open", "1. View service type description\n2. Check regulatory info", "Service type metadata and description shown", "Low"),
    ]),

    # ── Module 6: BPS Configuration ───────────────────────────────────────
    "BPSC": ("BPS Configuration", [
        ("View BPS config page", "Admin logged in", "1. Navigate to BPS Configuration", "BPS config page displayed with current settings", "High"),
        ("Update API key", "BPS config page open", "1. Enter new API key\n2. Save", "API key updated successfully", "High"),
        ("Set rate limit", "BPS config page open", "1. Enter rate limit value\n2. Save", "Rate limit configured", "Medium"),
        ("Health check - BPS service", "BPS config page open", "1. Click Health Check\n2. View result", "Health status displayed (up/down)", "High"),
        ("Add monitored province", "BPS config page open", "1. Click Add Area\n2. Select province\n3. Save", "Province added to monitored areas", "High"),
        ("Add monitored kabupaten", "Province selected", "1. Select kabupaten\n2. Save", "Kabupaten added to monitored areas", "High"),
        ("Remove monitored area", "Areas added", "1. Click remove on an area\n2. Confirm", "Area removed from monitoring", "Medium"),
        ("View monitored areas list", "Areas configured", "1. View areas list\n2. Verify count", "All monitored areas displayed", "High"),
        ("Invalid API key handling", "BPS config page open", "1. Enter invalid API key\n2. Test connection", "Error message displayed, key not saved", "Medium"),
        ("Config persistence after restart", "Config saved", "1. Save config\n2. Restart application\n3. Check config", "Config persisted correctly", "Medium"),
        ("BPS base URL configuration", "BPS config page open", "1. View/update base URL\n2. Save", "Base URL configured correctly", "Medium"),
        ("Rate limit enforcement", "Rate limit set", "1. Trigger many BPS API calls\n2. Exceed rate limit", "Rate limiting kicks in, requests queued", "Medium"),
        ("Config validation", "BPS config page open", "1. Enter invalid config values\n2. Try to save", "Validation errors shown", "Medium"),
        ("View API request history", "BPS calls made", "1. Navigate to API history\n2. View logs", "API request log displayed", "Low"),
        ("Multiple area bulk add", "BPS config page", "1. Select multiple provinces\n2. Add all at once", "All selected areas added", "Medium"),
        ("Area coverage indicator", "Areas configured", "1. View coverage percentage\n2. Check out of 34 provinces", "Coverage percentage displayed", "Low"),
        ("Config backup/export", "BPS config page", "1. Export current config\n2. Download", "Config exported for backup", "Low"),
        ("API quota monitoring", "BPS config active", "1. View API quota usage\n2. Check remaining calls", "API quota/usage displayed", "Medium"),
        ("Connection timeout handling", "BPS API slow", "1. Simulate slow response\n2. Verify timeout handling", "Timeout error shown after threshold", "Medium"),
        ("Config change history", "Config modified", "1. Modify config\n2. Check change log", "Config change logged with timestamp", "Low"),
        ("BPS config reset to default", "Config modified", "1. Click Reset to Default\n2. Confirm\n3. Verify values", "Config values reset to default settings", "Low"),
    ]),

    # ── Module 7: BPS Survey Management ───────────────────────────────────
    "BPSS": ("BPS Survey & Variable Management", [
        ("Search BPS catalog", "BPS config active", "1. Navigate to BPS Surveys\n2. Enter search keyword\n3. Search", "Matching BPS variables displayed", "High"),
        ("Register new variable", "Search results displayed", "1. Select variable from results\n2. Click Register\n3. Confirm", "Variable registered and active", "High"),
        ("View registered variables", "Variables registered", "1. Navigate to registered variables list", "All registered variables displayed with status", "High"),
        ("Toggle variable active/inactive", "Variable registered", "1. Click toggle on variable\n2. Verify status change", "Variable status toggled", "Medium"),
        ("Delete variable", "Variable registered", "1. Click delete on variable\n2. Confirm", "Variable removed from registered list", "Medium"),
        ("Filter variables by category", "Variables registered", "1. Select category filter\n2. Apply", "Variables filtered by category", "Medium"),
        ("Search within registered variables", "Variables registered", "1. Enter keyword in search\n2. View results", "Matching registered variables shown", "Medium"),
        ("Variable detail view", "Variable registered", "1. Click on variable\n2. View details", "Variable metadata displayed (unit, category, etc.)", "Medium"),
        ("Bulk register variables", "Search results with multiple items", "1. Select multiple variables\n2. Register all", "All selected variables registered", "Low"),
        ("Duplicate variable prevention", "Variable already registered", "1. Try to register same variable again", "Duplicate warning/prevention shown", "Medium"),
        ("Variable sync status", "Variables registered", "1. Check last sync date per variable", "Last sync timestamp displayed", "Low"),
        ("Search with no results", "BPS catalog search", "1. Search for non-existent keyword", "No results message displayed", "Low"),
        ("Category list population", "BPS surveys page", "1. Open category filter dropdown", "All available categories listed", "Low"),
        ("Variable count per category", "Categories displayed", "1. View category list\n2. Check count badges", "Correct count per category shown", "Low"),
        ("Variable data availability check", "Variable registered", "1. Check if data exists for variable\n2. View availability indicator", "Data availability status shown", "Medium"),
        ("Variable unit display", "Variable detail view", "1. View variable metadata\n2. Check unit information", "Unit of measurement displayed correctly", "Low"),
        ("Variable last sync timestamp", "Variable synced", "1. View last sync date per variable", "Last sync timestamp shown", "Low"),
        ("Bulk variable status toggle", "Multiple variables selected", "1. Select multiple variables\n2. Toggle status", "All selected variables toggled", "Low"),
        ("Variable export list", "Variables registered", "1. Click export\n2. Download list", "Variable list exported", "Low"),
    ]),

    # ── Module 8: BPS Data Visualization ──────────────────────────────────
    "BPSV": ("BPS Data Visualization", [
        ("View time-series line chart", "BPS data synced", "1. Navigate to BPS Visualization\n2. Select variable\n3. Select area\n4. View chart", "Line chart showing data over time", "High"),
        ("View bar chart comparison", "BPS data synced", "1. Select bar chart view\n2. Select multiple areas\n3. View chart", "Bar chart comparing areas", "High"),
        ("Filter by area", "BPS visualization open", "1. Select area from dropdown\n2. Apply", "Data filtered for selected area", "High"),
        ("Filter by variable", "BPS visualization open", "1. Select variable\n2. Apply", "Data displayed for selected variable", "High"),
        ("Filter by year range", "BPS visualization open", "1. Set start year\n2. Set end year\n3. Apply", "Data within year range displayed", "High"),
        ("Multiple area comparison", "BPS visualization open", "1. Select 2+ areas\n2. View comparison", "Multi-line chart with all areas", "Medium"),
        ("View sync history", "Sync operations completed", "1. Click sync history tab\n2. View log", "Sync operations listed with status", "Medium"),
        ("Trigger manual sync", "BPS config active", "1. Click 'Sync Now'\n2. Wait for completion", "Sync executes and history updated", "High"),
        ("Refresh data view", "Chart displayed", "1. Click Refresh\n2. Verify updated data", "Data refreshed from database", "Medium"),
        ("Export chart data", "Chart displayed", "1. Click export\n2. Download data", "Data exported to file", "Low"),
        ("Empty state - no data", "No BPS data synced", "1. Open visualization\n2. Select area with no data", "Empty state message shown", "Low"),
        ("Chart responsive resize", "Chart displayed", "1. Resize browser window\n2. Verify chart adjusts", "Chart resizes responsively", "Low"),
        ("Data tooltip on hover", "Chart displayed", "1. Hover over data point", "Tooltip shows exact value and date", "Medium"),
        ("Chart type toggle", "Chart displayed", "1. Toggle between line and bar\n2. Verify data consistency", "Same data displayed in different chart types", "Medium"),
        ("Sync status indicator", "BPS visualization open", "1. Check sync status badge\n2. Verify accuracy", "Current sync status shown (idle/syncing/error)", "Medium"),
        ("Full sync execution", "BPS configured", "1. Click Full Sync\n2. Monitor progress\n3. Verify completion", "All variables and areas synced", "High"),
        ("Incremental sync", "Previous full sync done", "1. Click Incremental Sync\n2. Verify only new data fetched", "Only new/updated data synced", "Medium"),
        ("Sync error handling", "BPS API unavailable", "1. Trigger sync when API is down\n2. Verify error handling", "Error message shown, graceful degradation", "Medium"),
        ("Data point annotation", "Chart with data", "1. Click on data point\n2. View annotation options", "Annotation/detail for data point shown", "Low"),
        ("Chart print/export", "Chart displayed", "1. Click print/export chart\n2. Download image", "Chart exported as image or PDF", "Low"),
        ("Multiple variable overlay", "BPS visualization", "1. Select 2+ variables\n2. View overlaid chart", "Multiple variables shown on same chart", "Medium"),
        ("Year range boundary validation", "Year filter", "1. Set start year > end year\n2. Verify validation", "Validation error for invalid range", "Medium"),
        ("Auto-refresh data view", "Visualization open", "1. Wait for auto-refresh interval\n2. Verify data updates", "Data auto-refreshes periodically", "Low"),
        ("Data comparison table", "BPS data available", "1. Select areas for comparison\n2. View table", "Side-by-side comparison table shown", "Medium"),
    ]),

    # ── Module 9: Telecom Potential ────────────────────────────────────────
    "TPOT": ("Telecom Market Potential", [
        ("View potential scores table", "Scores computed", "1. Navigate to Telecom Potential\n2. View scores table", "Area scores displayed with tiers (A/B/C/D)", "High"),
        ("View interactive map", "Scores computed", "1. Navigate to map view\n2. View choropleth", "Map displayed with color-coded areas", "High"),
        ("Click area on map", "Map displayed", "1. Click on a province/area\n2. View popup", "Area detail popup with scores shown", "High"),
        ("View radar chart", "Area selected", "1. Click area\n2. View radar chart", "Multi-dimensional radar chart displayed", "Medium"),
        ("View pie chart breakdown", "Area selected", "1. Click area\n2. View service breakdown", "Pie chart showing service distribution", "Medium"),
        ("Sort table by score", "Scores table displayed", "1. Click score column header\n2. Verify sort", "Table sorted by selected score", "Medium"),
        ("Filter by tier", "Scores displayed", "1. Select tier filter (A/B/C/D)\n2. Apply", "Only areas with selected tier shown", "Medium"),
        ("Configure scoring weights", "Admin logged in", "1. Open config\n2. Adjust weights\n3. Save", "Weights updated, ready for recompute", "Medium"),
        ("Recompute scores", "Config updated", "1. Click 'Compute Scores'\n2. Wait for completion", "Scores recomputed with new weights", "High"),
        ("V2 scoring with BPS data", "BPS data synced", "1. Enable V2 scoring\n2. Compute scores", "Scores include BPS demand component", "Medium"),
        ("Map zoom and pan", "Map displayed", "1. Zoom in/out\n2. Pan across map", "Map interaction smooth and responsive", "Medium"),
        ("Area detail dashboard", "Area clicked on map", "1. Click 'View Details'\n2. View area dashboard", "Full area dashboard with all metrics", "Medium"),
        ("Service breakdown per area", "Area detail open", "1. View service distribution\n2. Check counts", "Correct service counts per type shown", "Medium"),
        ("Export scores to Excel", "Scores displayed", "1. Click Export\n2. Download file", "Scores exported to Excel file", "Low"),
        ("GeoJSON data endpoint", "API available", "1. Request /v2/geojson\n2. Verify response", "Valid GeoJSON data returned", "Medium"),
        ("Sync pipeline execution", "Config set", "1. Click Full Sync Pipeline\n2. Monitor progress", "Fetch > Compute > Store pipeline executes", "High"),
        ("Year selection", "Multiple years available", "1. Select different year\n2. View scores", "Scores for selected year displayed", "Medium"),
        ("Market activity score detail", "Area selected", "1. View market activity breakdown\n2. Check components", "License count, diversity, and weighted scores shown", "Medium"),
        ("Opportunity score detail", "Area selected", "1. View untapped opportunity\n2. Check components", "Population potential and gap metrics shown", "Medium"),
        ("Config persistence", "Config saved", "1. Save config\n2. Reload page\n3. Verify", "Config values persisted correctly", "Medium"),
        ("Tier color coding accuracy", "Map displayed", "1. Verify color matches tier\n2. Check legend", "Colors correctly match A=green, B=blue, C=yellow, D=red", "Low"),
        ("No data handling", "New area with no licenses", "1. Compute scores for area with no data", "Zero scores handled gracefully, area shown as D tier", "Low"),
        ("Score comparison between years", "Multi-year data", "1. Select year A\n2. Select year B\n3. Compare scores", "Year-over-year score comparison shown", "Medium"),
        ("Map legend display", "Map rendered", "1. View map legend\n2. Check tier labels and colors", "Legend correctly shows tier color mapping", "Low"),
        ("Map full-screen mode", "Map displayed", "1. Click fullscreen button\n2. Verify map fills screen", "Map enters fullscreen mode", "Low"),
        ("Score weight validation", "Config editor open", "1. Enter weights totaling != 100%\n2. Save", "Validation error: weights must total 100%", "Medium"),
        ("Area ranking table export", "Scores computed", "1. Click export ranking\n2. Download", "Area ranking exported to Excel", "Low"),
        ("Score percentile calculation", "Scores computed", "1. View score percentiles\n2. Verify accuracy", "Percentile rankings shown correctly", "Medium"),
        ("Score trend analysis", "Multi-year scores", "1. Select area\n2. View score trend over time", "Score trend line chart displayed", "Medium"),
    ]),

    # ── Module 10: Dashboard ──────────────────────────────────────────────
    "DASH": ("Dashboard & Analytics", [
        ("View main dashboard", "User logged in", "1. Navigate to Dashboard\n2. View KPI cards", "Dashboard with service count cards displayed", "High"),
        ("KPI cards accuracy", "Data exists", "1. View KPI cards\n2. Compare with data count", "Card counts match actual data counts", "High"),
        ("Time-series trend chart", "Data exists over time", "1. View trend chart\n2. Verify data points", "Trend chart shows accurate data over time", "High"),
        ("Service distribution chart", "Data exists", "1. View distribution chart\n2. Verify segments", "Correct distribution of services shown", "Medium"),
        ("Dashboard auto-refresh", "Dashboard open", "1. Wait for auto-refresh interval\n2. Verify data update", "Dashboard data updates automatically", "Medium"),
        ("Dashboard responsive layout", "Dashboard open", "1. Resize browser\n2. Verify layout adapts", "Dashboard components reflow correctly", "Low"),
        ("Dashboard loading state", "Dashboard navigated to", "1. Navigate to dashboard\n2. Observe loading", "Loading skeleton/spinner shown while data loads", "Low"),
        ("Dashboard error state", "API unavailable", "1. Simulate API error\n2. View dashboard", "Error message displayed gracefully", "Medium"),
        ("Quick navigation from dashboard", "Dashboard displayed", "1. Click on a KPI card\n2. Verify navigation", "Navigates to relevant data view", "Medium"),
        ("Recent activity feed", "Actions performed", "1. View recent activity section\n2. Verify entries", "Recent actions displayed chronologically", "Medium"),
        ("Data freshness indicator", "Dashboard displayed", "1. Check last updated timestamp", "Last data refresh time displayed", "Low"),
        ("Dashboard permission-based content", "Different roles login", "1. Login as pelaku_usaha\n2. Login as admin\n3. Compare dashboards", "Content appropriate to role displayed", "Medium"),
        ("Enhanced dashboard analytics", "Data exists", "1. Navigate to data visualization\n2. View analytics", "Enhanced charts and analytics displayed", "Medium"),
        ("Chart interaction (tooltip/zoom)", "Chart displayed", "1. Hover over data points\n2. Try zoom/pan", "Tooltips shown, chart interactive", "Low"),
        ("Dashboard date range filter", "Dashboard displayed", "1. Set date range\n2. Apply filter\n3. Verify KPIs update", "KPIs recalculated for selected period", "Medium"),
        ("Dashboard export to PDF", "Dashboard displayed", "1. Click export\n2. Download PDF", "Dashboard exported as PDF report", "Low"),
        ("Dashboard widget arrangement", "Dashboard displayed", "1. View widget layout\n2. Verify logical grouping", "Widgets arranged logically by importance", "Low"),
        ("Dashboard data drill-down", "KPI card displayed", "1. Click on KPI card\n2. View detailed breakdown", "Drill-down shows detailed data behind KPI", "Medium"),
        ("Dashboard comparison mode", "Dashboard with data", "1. Enable comparison mode\n2. Select periods\n3. Compare", "Side-by-side period comparison shown", "Low"),
        ("Dashboard real-time update indicator", "Dashboard displayed", "1. Observe real-time indicator\n2. Verify pulse/animation", "Real-time update indicator visible", "Low"),
    ]),

    # ── Module 11: Support Ticketing ──────────────────────────────────────
    "TICK": ("Support & Ticketing", [
        ("Create new ticket", "User logged in", "1. Navigate to Support\n2. Fill ticket form\n3. Submit", "Ticket created with status 'open'", "High"),
        ("View ticket list", "Tickets exist", "1. Navigate to tickets\n2. View list", "Ticket list with status displayed", "High"),
        ("Send message on ticket", "Ticket open", "1. Open ticket\n2. Type message\n3. Send", "Message sent and displayed in conversation", "High"),
        ("Real-time message receive", "Ticket open, other party sending", "1. Keep ticket open\n2. Other party sends message\n3. Verify real-time update", "New message appears without refresh (WebSocket)", "High"),
        ("Close ticket", "Ticket open/in_progress", "1. Click Close Ticket\n2. Confirm", "Ticket status changed to 'closed'", "High"),
        ("Reopen ticket", "Ticket closed", "1. Open closed ticket\n2. Click Reopen", "Ticket status changed back to 'open'", "Medium"),
        ("Admin - view all tickets", "Admin logged in", "1. Navigate to Admin Tickets\n2. View all tickets", "All system tickets displayed", "High"),
        ("Admin - assign ticket", "Unassigned ticket exists", "1. Select ticket\n2. Assign to admin\n3. Save", "Ticket assigned, assignee updated", "High"),
        ("Admin - filter by status", "Admin tickets page", "1. Filter by status\n2. Apply", "Tickets filtered by status", "Medium"),
        ("Admin - search tickets", "Admin tickets page", "1. Enter search keyword\n2. Submit", "Matching tickets displayed", "Medium"),
        ("Unread message counter", "New messages received", "1. Check unread badge\n2. Verify count", "Correct unread count displayed", "High"),
        ("Mark messages as read", "Unread messages exist", "1. Open ticket\n2. Read messages", "Messages marked as read, counter decremented", "Medium"),
        ("Ticket status transitions", "Ticket exists", "1. Verify: open > in_progress > resolved > closed", "All valid transitions work", "High"),
        ("Ticket SLA tracking", "Ticket open for time", "1. Check SLA metrics\n2. Verify time tracking", "Response time and resolution time tracked", "Medium"),
        ("Ticket conversation threading", "Multiple messages", "1. View conversation\n2. Verify order", "Messages displayed in chronological order", "High"),
        ("File attachment in ticket", "Ticket open", "1. Attach file to message\n2. Send", "File uploaded and accessible", "Medium"),
        ("Ticket notification badge", "New ticket created", "1. Check notification area\n2. Verify badge", "New ticket notification shown", "Medium"),
        ("Empty ticket list state", "No tickets", "1. View tickets page with no tickets", "Empty state message displayed", "Low"),
        ("Concurrent messaging", "Two users on same ticket", "1. Both users send messages\n2. Verify sync", "Both messages appear in real-time for both users", "Medium"),
        ("Ticket auto-assignment", "New ticket, auto-assign enabled", "1. Create ticket\n2. Verify auto-assignment", "Ticket automatically assigned to available admin", "Low"),
        ("WebSocket reconnection", "Connection interrupted", "1. Disconnect network briefly\n2. Reconnect\n3. Verify sync", "WebSocket reconnects, messages synced", "Medium"),
        ("Ticket priority sorting", "Tickets with different priorities", "1. Sort by priority\n2. Verify order", "High priority tickets shown first", "Low"),
        ("Ticket attachment preview", "Ticket with attachment", "1. Click attachment\n2. Preview in modal", "Attachment previewed correctly", "Medium"),
        ("Ticket export", "Ticket list displayed", "1. Click Export\n2. Download", "Ticket list exported to Excel", "Low"),
        ("Ticket response time tracking", "Ticket with response", "1. View ticket metrics\n2. Check response time", "First response time tracked accurately", "Medium"),
        ("Ticket escalation", "Ticket open past SLA", "1. Check escalation rules\n2. Verify escalation trigger", "Ticket escalated after SLA breach", "Medium"),
        ("Ticket category/tag", "Ticket creation form", "1. Add category/tag to ticket\n2. Save\n3. Filter by tag", "Tickets categorizable and filterable", "Low"),
        ("Ticket satisfaction rating", "Ticket resolved", "1. Rate ticket resolution\n2. Submit rating", "Rating recorded for ticket", "Low"),
    ]),

    # ── Module 12: FAQ ────────────────────────────────────────────────────
    "FAQ": ("FAQ Management", [
        ("View public FAQ", "FAQs exist", "1. Navigate to /faq\n2. View FAQ list", "FAQs displayed with categories", "High"),
        ("Search FAQ", "FAQs exist", "1. Enter search keyword\n2. Submit", "Matching FAQs displayed", "High"),
        ("Filter FAQ by category", "FAQs with categories", "1. Select category\n2. Apply filter", "FAQs filtered by category", "Medium"),
        ("Admin - create FAQ", "Admin logged in", "1. Navigate to Admin FAQ\n2. Click Add\n3. Fill fields\n4. Save", "New FAQ created", "High"),
        ("Admin - edit FAQ", "FAQ exists", "1. Click Edit on FAQ\n2. Modify content\n3. Save", "FAQ updated successfully", "High"),
        ("Admin - delete FAQ", "FAQ exists", "1. Click Delete\n2. Confirm", "FAQ deleted", "High"),
        ("Admin - manage categories", "Admin FAQ page", "1. Add/edit/delete categories", "Categories managed successfully", "Medium"),
        ("FAQ accordion expand/collapse", "FAQ page open", "1. Click on FAQ question\n2. Verify answer expands\n3. Click again\n4. Verify collapse", "Accordion expands and collapses", "Medium"),
        ("FAQ ordering", "Multiple FAQs", "1. View FAQ order\n2. Verify logical ordering", "FAQs displayed in correct order", "Low"),
        ("FAQ empty state", "No FAQs exist", "1. View FAQ page with no entries", "Empty state message shown", "Low"),
        ("FAQ responsive layout", "Mobile viewport", "1. View FAQ on mobile", "FAQ renders correctly on mobile", "Low"),
        ("FAQ category count", "Categories with FAQs", "1. View category list\n2. Check count per category", "Correct FAQ count per category", "Low"),
        ("FAQ rich text content", "Admin FAQ editor", "1. Add formatted text\n2. Save\n3. View public", "Formatted content renders correctly", "Medium"),
        ("FAQ print view", "FAQ page open", "1. Click print\n2. Verify layout", "Print-friendly FAQ layout generated", "Low"),
        ("FAQ share link", "FAQ item open", "1. Copy share link\n2. Open in new tab", "Direct link opens specific FAQ item", "Low"),
        ("FAQ search no results", "FAQ page open", "1. Search for non-existent keyword", "No results message displayed", "Low"),
        ("FAQ multilingual content", "Admin FAQ editor", "1. Enter content in Bahasa\n2. Save\n3. View public", "Bahasa Indonesia content displays correctly", "Low"),
    ]),

    # ── Module 13: User Management ────────────────────────────────────────
    "USER": ("User Management", [
        ("View user list", "Admin logged in", "1. Navigate to User Management\n2. View user list", "All users displayed with roles", "High"),
        ("Search users", "Users exist", "1. Enter username/email in search\n2. Submit", "Matching users displayed", "High"),
        ("Edit user details", "User exists", "1. Click Edit on user\n2. Modify details\n3. Save", "User details updated", "High"),
        ("Assign role to user", "User exists", "1. Click on user\n2. Select role\n3. Assign", "Role assigned to user", "High"),
        ("Remove role from user", "User has assigned role", "1. Click on user\n2. Remove role\n3. Confirm", "Role removed from user", "High"),
        ("Filter users by role", "Users with different roles", "1. Select role filter\n2. Apply", "Users filtered by role", "Medium"),
        ("View user profile details", "User exists", "1. Click on user\n2. View profile", "Full user profile displayed", "Medium"),
        ("Prevent self-role-removal", "Admin viewing own profile", "1. Try to remove own admin role", "Action prevented with warning", "High"),
        ("User activity log view", "User has activity", "1. View user\n2. Check activity log", "User's recent activity displayed", "Medium"),
        ("User status toggle", "User exists", "1. Toggle active/inactive status\n2. Verify", "User status changed", "Medium"),
        ("User creation by admin", "Admin logged in", "1. Click Add User\n2. Fill details\n3. Save", "New user created with specified role", "Medium"),
        ("User pagination", "Many users exist", "1. Navigate pages\n2. Verify data", "User list paginates correctly", "Low"),
        ("User profile view by user", "User logged in", "1. Click Profile\n2. View own profile", "User can view own profile details", "High"),
        ("User profile update", "User logged in", "1. Edit own profile\n2. Save changes", "Profile updated successfully", "High"),
        ("User search by email", "Admin on user page", "1. Search by email address\n2. Verify results", "User found by email search", "Medium"),
        ("User role history", "User with role changes", "1. View user\n2. Check role change history", "Role change history displayed", "Low"),
        ("User last login display", "User list displayed", "1. View last login column\n2. Verify dates", "Last login timestamp shown per user", "Low"),
        ("Inactive user identification", "Users with no recent login", "1. Filter inactive users\n2. Verify criteria", "Users inactive > 90 days highlighted", "Low"),
        ("User export to Excel", "User list displayed", "1. Click Export\n2. Download", "User list exported to Excel", "Low"),
        ("User password reset by admin", "Admin on user detail", "1. Click Reset Password\n2. Confirm", "Password reset, user notified", "Medium"),
    ]),

    # ── Module 14: Permission Management ──────────────────────────────────
    "PERM": ("Permission Management", [
        ("View permission matrix", "Super admin logged in", "1. Navigate to Permissions\n2. View matrix", "Permission matrix displayed with roles and modules", "High"),
        ("Grant permission to role", "Permission matrix open", "1. Select role\n2. Select module\n3. Grant read/write", "Permission granted", "High"),
        ("Revoke permission", "Permission exists", "1. Select role\n2. Uncheck permission\n3. Save", "Permission revoked", "High"),
        ("Bulk permission update", "Permission matrix open", "1. Select multiple permissions\n2. Save all", "All permissions updated in batch", "Medium"),
        ("Permission enforcement - allowed", "User with read permission", "1. Login as user\n2. Access allowed page", "Page accessible and functional", "High"),
        ("Permission enforcement - denied", "User without permission", "1. Login as user\n2. Try to access restricted page", "Access denied, redirect or error shown", "High"),
        ("PermissionGuard component", "Protected component", "1. Render component with guard\n2. Login as authorized user", "Component renders for authorized user", "High"),
        ("PermissionGuard - unauthorized", "Protected component", "1. Render component\n2. Login as unauthorized user", "Component hidden or access denied", "High"),
        ("Module-level permissions", "Permissions configured", "1. Set read-only for module\n2. Login as user\n3. Try to edit", "Read allowed, write blocked", "High"),
        ("Field-level permissions", "Field permissions set", "1. Set field as hidden for role\n2. Login as that role", "Field not visible or editable", "Medium"),
        ("Permission template apply", "Template exists", "1. Select template\n2. Apply to role\n3. Verify permissions", "Template permissions applied to role", "Medium"),
        ("Permission change audit", "Permission changed", "1. Change permission\n2. Check audit log", "Permission change logged with details", "Medium"),
        ("API permission check", "API endpoint", "1. Call API without required permission\n2. Verify rejection", "401/403 response returned", "High"),
        ("Admin-only pages protection", "Non-admin user", "1. Try to access /users, /permissions\n2. Verify access denied", "Admin pages not accessible to non-admins", "High"),
        ("Permission export", "Permission matrix", "1. Export permissions\n2. Download", "Permission matrix exported", "Low"),
        ("Permission default reset", "Permissions modified", "1. Click Reset to Default\n2. Confirm", "Permissions reset to template defaults", "Low"),
        ("Permission conflict detection", "Overlapping rules", "1. Set conflicting permissions\n2. Verify handling", "Conflict detected and resolved (most restrictive wins)", "Medium"),
        ("Permission inheritance chain", "Nested permissions", "1. Set parent module permission\n2. Check child module\n3. Verify inheritance", "Child modules inherit parent permissions", "Medium"),
    ]),

    # ── Module 15: Integrations Dashboard ─────────────────────────────────
    "INTG": ("Integrations Dashboard", [
        ("View all integrations", "Integrations configured", "1. Navigate to Integrations\n2. View list", "All configured integrations displayed", "High"),
        ("View integration details", "Integration exists", "1. Click on integration\n2. View details", "Integration config and status shown", "High"),
        ("Toggle sync on/off", "Integration active", "1. Click toggle for integration\n2. Verify status change", "Sync enabled/disabled", "High"),
        ("View last sync status", "Sync completed", "1. View integration card\n2. Check status", "Last sync status (success/fail) and time shown", "High"),
        ("Trigger manual sync", "Integration active", "1. Click 'Sync Now'\n2. Wait for completion", "Manual sync executes, status updates", "High"),
        ("View sync error details", "Sync failed", "1. Click on failed sync\n2. View error", "Error message and details displayed", "Medium"),
        ("Monitor record count", "Sync completed", "1. View integration\n2. Check record count", "Total synced records count shown", "Medium"),
        ("Update integration config", "Integration exists", "1. Click Edit\n2. Modify config\n3. Save", "Integration config updated", "Medium"),
        ("API integration test dialog", "Integration dashboard", "1. Click 'Test Connection'\n2. View result", "Connection test result displayed", "Medium"),
        ("Integration health monitoring", "Integrations active", "1. View health status\n2. Check all integrations", "Health status (green/red) per integration", "Medium"),
        ("Sync history per integration", "Multiple syncs done", "1. Click on integration\n2. View history", "Chronological sync history displayed", "Medium"),
        ("KOMINFO tariff sync manual", "KOMINFO configured", "1. Navigate to KOMINFO sync\n2. Click Manual Sync\n3. Verify", "Tariff data synced from KOMINFO", "High"),
        ("KOMINFO sync status", "Sync in progress/done", "1. View sync status\n2. Verify accuracy", "Current sync status displayed", "High"),
        ("KOMINFO sync history", "Syncs completed", "1. View sync history\n2. Check entries", "All sync operations listed", "Medium"),
        ("KOMINFO synced data view", "Data synced", "1. Navigate to synced data\n2. View records", "KOMINFO tariff data displayed", "High"),
        ("e-Telekomunikasi connection test", "Config exists", "1. Test connection to e-Telekomunikasi\n2. View result", "Connection success/failure shown", "Medium"),
        ("Integration retry mechanism", "Sync failed", "1. View failed sync\n2. Click retry\n3. Verify retry", "Retry mechanism executes and updates status", "Medium"),
        ("Integration rate limit indicator", "Integration active", "1. View rate limit usage\n2. Check remaining quota", "Rate limit usage displayed per integration", "Low"),
        ("Integration data validation", "Data synced", "1. Validate synced data\n2. Check completeness", "Data integrity verified after sync", "Medium"),
        ("Integration schedule configuration", "Integration exists", "1. Set sync schedule\n2. Save\n3. Verify cron", "Sync schedule configured correctly", "Medium"),
    ]),

    # ── Module 16: Security & DevSecOps ───────────────────────────────────
    "SEC": ("Security & DevSecOps", [
        ("CORS enforcement", "API available", "1. Make cross-origin request from unlisted origin\n2. Verify rejection", "Request blocked by CORS policy", "High"),
        ("Helmet security headers", "API response received", "1. Inspect response headers\n2. Check CSP, HSTS, X-Frame", "All security headers present", "High"),
        ("Rate limiting on auth endpoints", "Login endpoint", "1. Send 100+ requests in 15 min\n2. Verify rate limit", "429 Too Many Requests returned", "High"),
        ("SQL injection prevention", "API with input", "1. Send SQL injection payload\n2. Verify handling", "Input sanitized, no SQL executed", "High"),
        ("XSS prevention", "Input field", "1. Enter <script> tag in input\n2. Submit\n3. Verify output", "Script tag sanitized, not executed", "High"),
        ("JWT token validation", "API endpoint", "1. Send request with invalid JWT\n2. Verify rejection", "401 Unauthorized returned", "High"),
        ("JWT token expiry enforcement", "Expired token", "1. Wait for token to expire\n2. Make request", "401 returned, token refresh required", "High"),
        ("File upload type validation", "Upload endpoint", "1. Try to upload .exe file\n2. Verify rejection", "Invalid file type rejected", "High"),
        ("File upload size validation", "Upload endpoint", "1. Try to upload >10MB file\n2. Verify rejection", "File size limit enforced", "Medium"),
        ("Audit log recording", "Any action performed", "1. Perform data action\n2. Check audit_logs\n3. Verify entry", "Action logged with timestamp, user, details", "High"),
        ("Login attempt tracking", "Login attempts made", "1. Make several login attempts\n2. Check login_attempts table", "All attempts recorded", "Medium"),
        ("Password hashing verification", "User registered", "1. Check password in database\n2. Verify hashed", "Password stored as bcrypt hash, not plaintext", "High"),
        ("Secure cookie flags", "Login response", "1. Inspect Set-Cookie header\n2. Check flags", "httpOnly, secure, sameSite flags set", "High"),
        ("HTTPS redirect", "HTTP request", "1. Make HTTP request\n2. Verify redirect", "Redirected to HTTPS", "Medium"),
        ("API error handling (no stack traces)", "API error triggered", "1. Trigger server error\n2. Check response", "Generic error message, no stack trace in production", "Medium"),
        ("DevSecOps monitoring dashboard", "Admin logged in", "1. Navigate to security dashboard\n2. View metrics", "Security metrics and audit logs displayed", "Medium"),
        ("Session management security", "User session active", "1. Verify session isolation\n2. Check no session hijacking possible", "Sessions properly isolated", "High"),
        ("Access control on admin routes", "Non-admin user", "1. Try to access admin API endpoints\n2. Verify rejection", "403 Forbidden returned", "High"),
        ("OWASP A01 Broken Access Control", "API endpoints", "1. Test IDOR on /telekom-data/:id\n2. Access other user's data", "Access denied for unauthorized resources", "High"),
        ("OWASP A03 Injection", "Input fields", "1. Test NoSQL injection\n2. Test command injection\n3. Verify prevention", "All injection attempts prevented", "High"),
        ("OWASP A07 Auth Failures", "Auth endpoints", "1. Test credential stuffing\n2. Test default passwords\n3. Verify prevention", "Brute force and credential attacks prevented", "High"),
        ("Security header X-Content-Type-Options", "API response", "1. Check X-Content-Type-Options header\n2. Verify nosniff", "Header present with nosniff value", "Medium"),
        ("Security header X-Frame-Options", "API response", "1. Check X-Frame-Options header\n2. Verify DENY/SAMEORIGIN", "Header present preventing clickjacking", "Medium"),
        ("API key rotation", "Service key in use", "1. Generate new API key\n2. Update config\n3. Verify old key rejected", "Old key rejected, new key works", "Medium"),
    ]),

    # ── Module 17: RBAC Enforcement ───────────────────────────────────────
    "RBAC": ("RBAC Enforcement", [
        ("Super admin - full access", "super_admin logged in", "1. Access all pages\n2. Access all APIs", "All pages and APIs accessible", "High"),
        ("Internal admin - admin access", "internal_admin logged in", "1. Access admin pages\n2. Verify no user management", "Admin pages accessible, user mgmt restricted", "High"),
        ("Pelaku usaha - limited access", "pelaku_usaha logged in", "1. Access dashboard\n2. Try admin pages\n3. Verify restrictions", "Dashboard accessible, admin pages blocked", "High"),
        ("Pengolah data - data access", "pengolah_data logged in", "1. Access data management\n2. Access analytics\n3. Try admin pages", "Data/analytics accessible, admin blocked", "High"),
        ("Internal group - read access", "internal_group logged in", "1. Access data (read)\n2. Try to edit\n3. Verify restrictions", "Read access granted, write blocked", "High"),
        ("Guest - public only", "guest role", "1. Access public pages\n2. Try protected pages\n3. Verify restrictions", "Only public pages accessible", "High"),
        ("Role-based sidebar menu", "Different roles login", "1. Login as each role\n2. Compare sidebar menus", "Menu items match role permissions", "High"),
        ("API permission by role", "Different roles", "1. Call admin API as pelaku_usaha\n2. Verify rejection", "Permission denied for unauthorized role", "High"),
        ("Role assignment persistence", "Role assigned", "1. Assign role\n2. Logout\n3. Login again\n4. Verify role", "Role persists across sessions", "High"),
        ("Multiple role handling", "User with multiple roles", "1. Assign multiple roles\n2. Verify combined permissions", "User gets union of all role permissions", "Medium"),
        ("Role-based data filtering", "Data with ownership", "1. Login as pelaku_usaha\n2. View data\n3. Verify only own data shown", "Only user's company data visible", "High"),
        ("Middleware access control", "Protected route", "1. Send API request without proper role\n2. Check middleware rejection", "Access control middleware blocks request", "High"),
        ("Role hierarchy enforcement", "Different roles", "1. Verify super_admin > internal_admin > pelaku_usaha\n2. Check inheritance", "Role hierarchy correctly enforced", "High"),
        ("API data scoping by role", "pelaku_usaha logged in", "1. Call /telekom-data API\n2. Verify only company data returned", "Data scoped to user's company", "High"),
        ("Unauthorized API call logging", "Unauthorized request", "1. Make unauthorized API call\n2. Check audit log", "Unauthorized attempt logged for security review", "Medium"),
        ("Role change immediate effect", "User role changed", "1. Admin changes user's role\n2. User's next request\n3. Verify new permissions", "New permissions take effect immediately", "High"),
    ]),

    # ── Module 18: Navigation & Layout ────────────────────────────────────
    "NAV": ("Navigation & Layout", [
        ("Sidebar navigation", "User logged in", "1. Click sidebar items\n2. Verify navigation", "All sidebar links navigate correctly", "High"),
        ("Sidebar collapse/expand", "Sidebar visible", "1. Click collapse button\n2. Verify sidebar collapses\n3. Click expand", "Sidebar toggles correctly", "Medium"),
        ("Breadcrumb navigation", "Deep page open", "1. View breadcrumb\n2. Click parent", "Breadcrumb shows path, clicking navigates", "Medium"),
        ("404 Not Found page", "Invalid URL", "1. Navigate to invalid URL\n2. View result", "404 page displayed", "Medium"),
        ("Auth redirect - unauthenticated", "Not logged in", "1. Try to access protected page\n2. Verify redirect", "Redirected to login page", "High"),
        ("Auth redirect - after login", "Redirected to login", "1. Login successfully\n2. Verify redirect back", "Redirected to originally requested page", "Medium"),
        ("Page title updates", "Navigate between pages", "1. Navigate to different pages\n2. Check browser tab title", "Page title updates correctly", "Low"),
        ("Loading states", "Page loading", "1. Navigate to data-heavy page\n2. Observe loading", "Loading indicator shown during data fetch", "Medium"),
        ("Error boundary", "Component error", "1. Trigger component error\n2. Verify error boundary", "Error boundary catches and displays fallback", "Medium"),
        ("Toast notifications", "Action performed", "1. Perform action (save, delete)\n2. Check toast", "Toast notification shown with result", "Medium"),
        ("Mobile hamburger menu", "Mobile viewport", "1. View on mobile\n2. Click hamburger\n3. Navigate", "Mobile menu opens and navigates", "Low"),
        ("Dark mode toggle", "Theme toggle available", "1. Toggle dark mode\n2. Verify theme change", "UI switches between light and dark themes", "Low"),
        ("Responsive table layout", "Table displayed on mobile", "1. View data table on small screen", "Table scrolls horizontally or adapts", "Low"),
        ("Keyboard navigation", "Any page", "1. Use Tab to navigate\n2. Use Enter to activate", "All interactive elements keyboard accessible", "Low"),
        ("URL deep linking", "Specific page URL", "1. Copy URL of a page\n2. Open in new browser\n3. Verify navigation", "Page loads correctly from deep link", "Medium"),
        ("Browser refresh state persistence", "Page with filters", "1. Apply filters\n2. Refresh browser\n3. Check state", "Filters persist after refresh (or clear gracefully)", "Medium"),
        ("Multi-tab navigation", "App open in multiple tabs", "1. Open app in 2 tabs\n2. Login in tab 1\n3. Navigate in tab 2", "Session shared across tabs correctly", "Medium"),
        ("Scroll to top on navigation", "Long page scrolled down", "1. Scroll down\n2. Navigate to new page", "New page starts at top", "Low"),
        ("Active menu item highlight", "Any page open", "1. Navigate to a page\n2. Check sidebar", "Current page highlighted in sidebar", "Low"),
        ("Page transition animation", "Navigation between pages", "1. Navigate between pages\n2. Observe transition", "Smooth page transition without flicker", "Low"),
        ("Footer content display", "Any page", "1. Scroll to bottom\n2. View footer", "Footer shows copyright, version, contact info", "Low"),
    ]),

    # ── Module 19: Data Export & Import ───────────────────────────────────
    "EXIM": ("Data Export & Import", [
        ("Excel export - full dataset", "Data available", "1. Navigate to data page\n2. Click Export\n3. Download", "Excel file with all data downloaded", "High"),
        ("Excel export - filtered data", "Filters applied", "1. Apply filters\n2. Click Export\n3. Download", "Only filtered data in export", "High"),
        ("Excel import - valid file", "Import dialog open", "1. Upload valid Excel\n2. Map columns\n3. Import", "Data imported successfully", "High"),
        ("Excel import - column mapper", "File uploaded", "1. View column mapper\n2. Map source to target\n3. Verify preview", "Columns correctly mapped with preview", "High"),
        ("Excel import - data preview", "Columns mapped", "1. Review preview\n2. Check data accuracy\n3. Confirm", "Data preview accurate before import", "High"),
        ("Excel import - validation errors", "Invalid data in file", "1. Upload file with invalid data\n2. Verify validation", "Validation errors shown per row/column", "High"),
        ("Excel import - duplicate detection", "Data already exists", "1. Import file with existing records\n2. Verify handling", "Duplicates detected and handled (skip/update)", "Medium"),
        ("Export with formatting", "Data exported", "1. Open exported Excel\n2. Check formatting", "Excel has proper headers, formatting", "Medium"),
        ("Large file import (1000+ rows)", "Large Excel file", "1. Upload large file\n2. Import\n3. Verify all rows", "All rows imported, progress indicator shown", "Medium"),
        ("Import cancel", "Import in progress", "1. Start import\n2. Cancel mid-process", "Import cancelled, no partial data", "Medium"),
        ("PDF document preview", "Document uploaded", "1. Click on uploaded document\n2. View preview", "PDF opens in preview modal", "Medium"),
        ("File download", "Documents attached", "1. Click download link\n2. Verify file", "File downloads correctly", "Medium"),
        ("Excel template download", "Import dialog open", "1. Click 'Download Template'\n2. Open template", "Template Excel with correct columns downloaded", "Medium"),
        ("Import progress indicator", "Large file importing", "1. Start large import\n2. Observe progress", "Progress bar/percentage shown during import", "Medium"),
        ("Export filename convention", "Export triggered", "1. Download export\n2. Check filename", "Filename includes date and module name", "Low"),
        ("Import error log export", "Import with errors", "1. Import file with errors\n2. Export error log", "Error log downloaded with row/error details", "Medium"),
        ("Multi-sheet Excel import", "Excel with multiple sheets", "1. Upload multi-sheet file\n2. Select sheet\n3. Import", "User can select target sheet for import", "Medium"),
        ("Data format auto-detection", "Excel upload", "1. Upload Excel with dates\n2. Check date format detection", "Date formats auto-detected and converted", "Low"),
    ]),

    # ── Module 20: Location Management ────────────────────────────────────
    "LOC": ("Location Management", [
        ("Province selector", "Any form with location", "1. Open province dropdown\n2. View provinces", "All 34 provinces listed", "High"),
        ("Kabupaten selector", "Province selected", "1. Select province\n2. Open kabupaten dropdown", "Kabupaten/kota for selected province listed", "High"),
        ("Province-kabupaten dependency", "Location selector", "1. Select province\n2. Change province\n3. Verify kabupaten resets", "Kabupaten list updates when province changes", "High"),
        ("Location data completeness", "Location selectors", "1. Check all provinces\n2. Spot-check kabupaten lists", "All provinces and kabupaten present", "Medium"),
        ("Location-based filtering", "Data with locations", "1. Filter by province\n2. Filter by kabupaten\n3. Verify results", "Data correctly filtered by location", "High"),
        ("Location in data records", "Create/edit record", "1. Set location for record\n2. Save\n3. Verify", "Location saved and displayed correctly", "High"),
        ("Location search", "Location selector", "1. Type in location search\n2. Verify suggestions", "Matching locations suggested", "Medium"),
        ("Location migration tool", "Migration needed", "1. Run location migration\n2. Verify data updated", "Location data migrated correctly", "Low"),
        ("Multiple location records", "Data management", "1. Create records with different locations\n2. Verify each", "All locations saved correctly", "Medium"),
        ("Location display format", "Record with location", "1. View record detail\n2. Check location display", "Location shows province - kabupaten format", "Low"),
        ("Location data integrity check", "Location selectors", "1. Check parent-child integrity\n2. Verify no orphan kabupaten", "All kabupaten have valid parent province", "Medium"),
        ("Location code mapping", "Location with codes", "1. View location code\n2. Verify matches BPS code", "Location codes match BPS standard codes", "Medium"),
        ("Location autocomplete speed", "Location selector", "1. Type location name\n2. Measure suggestion speed", "Suggestions appear within 300ms", "Low"),
    ]),

    # ── Module 21: Real-time Features ─────────────────────────────────────
    "RT": ("Real-time & WebSocket", [
        ("WebSocket connection establish", "User logged in", "1. Login to system\n2. Verify WS connection", "WebSocket connection established", "High"),
        ("Real-time ticket message", "Ticket open on two clients", "1. Send message from client A\n2. Verify appears on client B", "Message appears in real-time", "High"),
        ("Unread count real-time update", "New message received", "1. Receive message while on different page\n2. Check unread badge", "Badge count increments in real-time", "High"),
        ("WebSocket reconnection", "Connection dropped", "1. Simulate network disconnect\n2. Reconnect\n3. Verify sync", "Connection re-established, data synced", "High"),
        ("Data broadcast on change", "Data modified", "1. Modify data from one session\n2. Check another session", "Data change reflected on other sessions", "Medium"),
        ("WebSocket authentication", "Unauthenticated client", "1. Try WS connection without auth\n2. Verify rejection", "Unauthenticated WS connection rejected", "High"),
        ("Multiple concurrent connections", "Multiple users online", "1. Connect 10+ users\n2. Send messages\n3. Verify all receive", "All concurrent users receive messages", "Medium"),
        ("WebSocket message ordering", "Rapid messages sent", "1. Send multiple messages quickly\n2. Verify order", "Messages arrive in correct order", "Medium"),
        ("Connection status indicator", "WebSocket active", "1. View connection status\n2. Disconnect\n3. Verify indicator", "Status shows connected/disconnected", "Medium"),
        ("Graceful disconnection", "User logs out", "1. Logout\n2. Verify WS cleanup", "WebSocket connection properly closed", "Medium"),
        ("WebSocket heartbeat/ping", "Connection established", "1. Monitor WS connection\n2. Check heartbeat interval", "Heartbeat keeps connection alive", "Medium"),
        ("WebSocket message size limit", "Large message", "1. Send very large message via WS\n2. Verify handling", "Large messages handled or rejected gracefully", "Low"),
        ("WebSocket cross-tab sync", "App in multiple tabs", "1. Receive message in tab 1\n2. Check tab 2", "Message state synced across browser tabs", "Medium"),
    ]),

    # ── Module 22: Performance ────────────────────────────────────────────
    "PERF": ("Performance & Load", [
        ("Page load time < 3s", "Any page", "1. Navigate to page\n2. Measure load time", "Page loads within 3 seconds", "High"),
        ("API response < 500ms", "API endpoint", "1. Make API request\n2. Measure response time", "Response within 500ms (p95)", "High"),
        ("Pagination performance", "Large dataset (1000+)", "1. Navigate pages\n2. Measure load time", "Each page loads within acceptable time", "High"),
        ("Search performance", "Large dataset", "1. Execute search\n2. Measure response", "Search results return within 1 second", "Medium"),
        ("Map rendering performance", "Telecom potential map", "1. Load map\n2. Measure render time", "Map renders within 3 seconds", "Medium"),
        ("Chart rendering performance", "Dashboard with charts", "1. Load charts\n2. Measure render time", "Charts render within 2 seconds", "Medium"),
        ("Excel export performance", "Large dataset", "1. Export 1000+ records\n2. Measure time", "Export completes within 10 seconds", "Medium"),
        ("Excel import performance", "Large file", "1. Import 1000+ rows\n2. Measure time", "Import processes within 30 seconds", "Medium"),
        ("Concurrent user handling", "Multiple users", "1. Simulate 50+ concurrent users\n2. Monitor performance", "System remains responsive", "Medium"),
        ("Database query optimization", "Complex queries", "1. Execute complex query\n2. Check query time", "Query time < 200ms with proper indexing", "Medium"),
        ("Memory usage monitoring", "Extended session", "1. Use system for extended period\n2. Monitor memory", "No memory leaks, stable usage", "Low"),
        ("Network bandwidth optimization", "Page load", "1. Measure transferred data\n2. Check compression", "Responses gzipped, assets minified", "Low"),
        ("Lazy loading verification", "Page with many components", "1. Open page\n2. Check network tab for lazy chunks", "Components lazy loaded on demand", "Low"),
        ("Cache invalidation", "Data updated", "1. Update data\n2. Verify cache updated\n3. Check stale data", "React Query cache invalidated on mutation", "Medium"),
        ("Image optimization", "Page with images/icons", "1. Check image sizes\n2. Verify format optimization", "Images optimized for web delivery", "Low"),
        ("Bundle size analysis", "Production build", "1. Run build\n2. Analyze bundle size", "Bundle size within acceptable limits", "Low"),
    ]),

    # ── Module 23: Error Handling ─────────────────────────────────────────
    "ERR": ("Error Handling & Edge Cases", [
        ("Network offline handling", "Connection lost", "1. Disconnect network\n2. Try to navigate\n3. Reconnect", "Offline state handled, reconnection seamless", "Medium"),
        ("API timeout handling", "Slow API response", "1. Simulate slow API\n2. Wait for timeout", "Timeout error displayed to user", "Medium"),
        ("Invalid route handling", "Invalid URL", "1. Enter random URL path\n2. Verify result", "404 page displayed", "High"),
        ("Form validation - all fields", "Any form", "1. Submit empty form\n2. Verify all validations", "All required field errors shown", "High"),
        ("Server error handling (500)", "API error", "1. Trigger server error\n2. Check UI response", "User-friendly error message shown", "High"),
        ("Session expiry mid-action", "Token expires during use", "1. Work in system\n2. Token expires\n3. Try action", "Graceful redirect to login or token refresh", "High"),
        ("Concurrent edit conflict", "Same record edited by two users", "1. User A edits record\n2. User B edits same\n3. Both save", "Conflict detected, latest save wins or warning shown", "Medium"),
        ("Database connection loss", "DB goes down", "1. Simulate DB disconnect\n2. Make request", "Error message shown, no crash", "High"),
        ("File upload failure recovery", "Upload interrupted", "1. Start file upload\n2. Interrupt upload", "Upload failure handled, user can retry", "Medium"),
        ("Invalid data input handling", "Data form", "1. Enter special characters\n2. Enter very long strings\n3. Submit", "Data validated and sanitized", "Medium"),
        ("Browser back/forward navigation", "Multi-step flow", "1. Use browser back\n2. Use browser forward", "Navigation works correctly without breaking state", "Medium"),
        ("Duplicate submission prevention", "Form submission", "1. Click submit rapidly\n2. Verify only one submission", "Double submission prevented", "Medium"),
        ("Unicode/special char handling", "Input fields", "1. Enter Unicode characters (emoji, CJK)\n2. Save\n3. Verify display", "Unicode characters stored and displayed correctly", "Medium"),
        ("Large text input handling", "Textarea fields", "1. Enter very long text (10000+ chars)\n2. Save", "Long text handled without truncation or error", "Medium"),
        ("API 404 handling", "Non-existent resource", "1. Request /telekom-data/99999\n2. Check response", "404 Not Found with appropriate message", "Medium"),
        ("API 422 validation error", "Invalid request body", "1. Send malformed request body\n2. Check response", "422 with field-level validation errors", "Medium"),
        ("Timezone handling", "Dates in system", "1. Check date display\n2. Verify timezone consistency", "All dates display in consistent timezone (WIB)", "Medium"),
        ("Pagination boundary - first page", "First page of data", "1. Click Previous on first page", "Previous button disabled on first page", "Low"),
        ("Pagination boundary - last page", "Last page of data", "1. Click Next on last page", "Next button disabled on last page", "Low"),
        ("Empty database handling", "Fresh database", "1. Login with no data\n2. Navigate all pages", "All pages handle empty state gracefully", "Medium"),
        ("Special chars in search", "Search field", "1. Enter special characters in search\n2. Execute search", "Search handles special chars without error", "Medium"),
        ("Bulk operation progress", "Bulk action initiated", "1. Start bulk operation\n2. Monitor progress", "Progress indicator shows completion status", "Low"),
    ]),

    # ── Module 24: Deployment & Health ────────────────────────────────────
    "DEPL": ("Deployment & Health Check", [
        ("Health check endpoint", "Server running", "1. GET /health\n2. Verify response", "200 OK with health status", "High"),
        ("Server startup verification", "Fresh start", "1. Start server\n2. Verify all services", "Server starts, all routes registered", "High"),
        ("Database connection on startup", "Server starting", "1. Start server\n2. Verify DB connection", "Database connected, tables initialized", "High"),
        ("Environment variable loading", "Server starting", "1. Set env vars\n2. Start server\n3. Verify config", "All env vars loaded correctly", "High"),
        ("Frontend build verification", "npm run build", "1. Run build\n2. Verify output", "Build succeeds, assets generated in dist/", "High"),
        ("Frontend serving at /panel/", "Built frontend", "1. Access /panel/\n2. Verify SPA loads", "Frontend loads at correct base path", "High"),
        ("API proxy configuration", "Dev environment", "1. Start dev server\n2. Make API call\n3. Verify proxy", "Frontend proxy to backend works", "Medium"),
        ("Static file serving", "Server running", "1. Request static asset\n2. Verify delivery", "Static files served correctly", "Medium"),
        ("WebSocket server startup", "Server starting", "1. Start server\n2. Verify WS endpoint", "WebSocket server listening on port 4000", "High"),
        ("Docker container build", "Dockerfile present", "1. Build Docker image\n2. Verify success", "Docker image builds successfully", "Medium"),
        ("PM2 process management", "PM2 configured", "1. Start with PM2\n2. Verify process running", "Process running, auto-restart configured", "Medium"),
        ("Graceful shutdown", "Server running", "1. Send SIGTERM\n2. Verify cleanup", "Server shuts down gracefully, connections closed", "Medium"),
        ("Database migration on startup", "New migrations", "1. Start with pending migrations\n2. Verify execution", "Migrations run automatically", "Medium"),
        ("Log file creation", "Server running", "1. Check log output\n2. Verify format", "Logs written with proper format and level", "Low"),
        ("CORS config verification", "Server running", "1. Send request from allowed origin\n2. Send from disallowed", "Allowed origins accepted, others rejected", "High"),
        ("Backup and restore verification", "Database with data", "1. Create backup\n2. Restore backup\n3. Verify data", "Data fully restored from backup", "Medium"),
        ("SSL certificate verification", "HTTPS enabled", "1. Check SSL certificate\n2. Verify validity", "Valid SSL certificate installed", "High"),
        ("Server restart recovery", "Server crashed/restarted", "1. Kill server process\n2. PM2 auto-restarts\n3. Verify service", "Server auto-restarts, no data loss", "High"),
        ("Concurrent deployment", "Deployment in progress", "1. Deploy new version\n2. Verify zero downtime", "Zero downtime deployment achieved", "Medium"),
        ("Environment-specific config", "Different environments", "1. Check dev config\n2. Check prod config\n3. Verify isolation", "Environment configs properly isolated", "Medium"),
        ("Disk space monitoring", "Server running", "1. Check upload storage space\n2. Verify threshold alerts", "Low disk space alerts configured", "Low"),
        ("Server timezone configuration", "Server running", "1. Check server timezone\n2. Verify matches WIB (Asia/Jakarta)", "Server timezone set to Asia/Jakarta", "Medium"),
    ]),
}

# ── Generate all sheets ───────────────────────────────────────────────────
total_scenarios = 0
first_sheet = True
for sheet_key, (module_name, scenarios) in ALL_MODULES.items():
    if first_sheet:
        ws = wb.active
        first_sheet = False
    else:
        ws = wb.create_sheet()
    count = create_sheet(ws, sheet_key, module_name, scenarios, 1)
    total_scenarios += count
    print(f"  {sheet_key}: {count} scenarios")

# ── Summary Sheet ─────────────────────────────────────────────────────────
ws_summary = wb.create_sheet("SUMMARY", 0)
ws_summary.merge_cells('A1:F1')
cell = ws_summary['A1']
cell.value = "UAT SCENARIOS SUMMARY — TELKOM INSIGHT HUB"
cell.font = Font(name="Calibri", size=16, bold=True, color="E11A27")
cell.alignment = Alignment(horizontal='center')

ws_summary.merge_cells('A2:F2')
cell = ws_summary['A2']
cell.value = "Kementerian Komunikasi dan Digital — Direktorat Jenderal PPI"
cell.font = Font(name="Calibri", size=11, color="666666")
cell.alignment = Alignment(horizontal='center')

# Module summary table
headers = ["#", "Module Code", "Module Name", "Scenarios", "Priority High", "Sheet"]
for ci, (header, width) in enumerate(zip(headers, [5, 14, 35, 12, 12, 12]), 1):
    cell = ws_summary.cell(row=4, column=ci, value=header)
    cell.font = HEADER_FONT
    cell.fill = RED_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER
    ws_summary.column_dimensions[get_column_letter(ci)].width = width

for i, (sheet_key, (module_name, scenarios)) in enumerate(ALL_MODULES.items()):
    row = 5 + i
    high_count = sum(1 for s in scenarios if s[4] == "High")
    row_data = [i + 1, sheet_key, module_name, len(scenarios), high_count, sheet_key]
    for ci, val in enumerate(row_data, 1):
        cell = ws_summary.cell(row=row, column=ci, value=val)
        cell.font = NORMAL_FONT
        cell.alignment = CENTER if ci != 3 else LEFT_WRAP
        cell.border = THIN_BORDER
        if i % 2 == 1:
            cell.fill = LIGHT_GRAY_FILL

# Total row
total_row = 5 + len(ALL_MODULES)
ws_summary.cell(row=total_row, column=1).value = ""
ws_summary.merge_cells(f'B{total_row}:C{total_row}')
cell = ws_summary.cell(row=total_row, column=2, value="TOTAL")
cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
cell.fill = DARK_FILL
cell.alignment = CENTER
cell.border = THIN_BORDER
ws_summary.cell(row=total_row, column=3).fill = DARK_FILL
ws_summary.cell(row=total_row, column=3).border = THIN_BORDER

cell = ws_summary.cell(row=total_row, column=4, value=total_scenarios)
cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
cell.fill = DARK_FILL
cell.alignment = CENTER
cell.border = THIN_BORDER

high_total = sum(sum(1 for s in scenarios if s[4] == "High") for _, (_, scenarios) in ALL_MODULES.items())
cell = ws_summary.cell(row=total_row, column=5, value=high_total)
cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
cell.fill = DARK_FILL
cell.alignment = CENTER
cell.border = THIN_BORDER

ws_summary.cell(row=total_row, column=6).fill = DARK_FILL
ws_summary.cell(row=total_row, column=6).border = THIN_BORDER

# Sign-off on summary
sr = total_row + 3
ws_summary.merge_cells(f'A{sr}:F{sr}')
cell = ws_summary.cell(row=sr, column=1, value="DOCUMENT SIGN-OFF / PERSETUJUAN DOKUMEN")
cell.font = Font(name="Calibri", size=14, bold=True, color="E11A27")
cell.alignment = Alignment(horizontal='center')

sr += 1
signoff_headers = ["Peran", "Nama", "Jabatan", "Tanda Tangan", "Tanggal", ""]
for ci, header in enumerate(signoff_headers, 1):
    cell = ws_summary.cell(row=sr, column=ci, value=header)
    cell.font = WHITE_FONT
    cell.fill = DARK_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER

signoff_roles = [
    "Disusun oleh (Prepared by)",
    "QA Lead (Test Lead)",
    "Project Manager (PM)",
    "Teknis Lead (Technical Approval)",
    "Stakeholder (Business Approval)",
    "Management (Final Sign-off)",
]

for ri, role in enumerate(signoff_roles):
    row = sr + 1 + ri
    ws_summary.cell(row=row, column=1, value=role).font = SIGNOFF_FONT
    ws_summary.cell(row=row, column=1).fill = SIGNOFF_FILL
    ws_summary.cell(row=row, column=1).border = THIN_BORDER
    ws_summary.cell(row=row, column=1).alignment = LEFT_WRAP
    for ci in range(2, 7):
        cell = ws_summary.cell(row=row, column=ci)
        cell.fill = SIGNOFF_FILL
        cell.border = THIN_BORDER
        cell.alignment = CENTER
    ws_summary.row_dimensions[row].height = 45

# Info box
info_row = sr + len(signoff_roles) + 3
ws_summary.merge_cells(f'A{info_row}:F{info_row}')
cell = ws_summary.cell(row=info_row, column=1,
                        value="Dokumen UAT ini berisi 528 skenario pengujian pada 24 modul. Seluruh skenario harus dieksekusi dan ditandatangani sebelum sistem dapat dinyatakan Production Ready.")
cell.font = Font(name="Calibri", size=10, italic=True, color="666666")
cell.alignment = Alignment(horizontal='center', wrap_text=True)

info_row += 1
ws_summary.merge_cells(f'A{info_row}:F{info_row}')
cell = ws_summary.cell(row=info_row, column=1,
                        value=f"Nomor Dokumen: TIH-UAT-2026-001  |  Versi: 1.0  |  Tanggal: April 2026  |  Total Scenarios: {total_scenarios}")
cell.font = Font(name="Calibri", size=10, color="999999")
cell.alignment = Alignment(horizontal='center')

# ── Save ──────────────────────────────────────────────────────────────────
path = os.path.join(OUTPUT_DIR, "UAT_SCENARIOS_SIGNED.xlsx")
wb.save(path)
print(f"\nSaved: {path}")
print(f"Total scenarios: {total_scenarios}")
print(f"Total sheets: {len(ALL_MODULES) + 1} (including summary)")
